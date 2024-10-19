import json
import pandas as pd 
import requests
from bs4 import BeautifulSoup
import openpyxl
from fastapi import FastAPI
app = FastAPI()
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from fastapi.middleware.cors import CORSMiddleware

origins = [
    "http://localhost:5173",  # El puerto 3000 es común para aplicaciones React
    "http://127.0.0.1:5173",  # También podrías usar 127.0.0.1
    "http://localhost:8080",  # Otros puertos que quieras permitir
]

# Añadir middleware de CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,  # Permitir estos orígenes
    allow_credentials=True,  # Permitir el envío de cookies con la solicitud
    allow_methods=["*"],  # Permitir todos los métodos (GET, POST, etc.)
    allow_headers=["*"],  # Permitir todos los encabezados
)

@app.get("/search-api")
def search(key : str): #keyword: None
  

    driver = webdriver.Edge(service=Service(EdgeChromiumDriverManager().install()))

    # Creamos un nuevo libro de trabajo de Excel
    workbook = openpyxl.Workbook()

    # Seleccionamos la hoja activa
    hoja_activa = workbook.active


    especie = []
    nombre = []
    descarga = []

    # URL inicial
    ncbi = 'https://www.ncbi.nlm.nih.gov/gds/?term=' + key
    driver.get(ncbi)

    contador_paginas = 0

    #while True:
    while contador_paginas < 4:
        time.sleep(2)  # Esperar a que la página cargue completamente

        # Obtener el contenido de la página actual
        ncbisoup = BeautifulSoup(driver.page_source, 'html.parser')

        # Encuentra todos los elementos div con la clase 'details lefty'
        elementos_a = ncbisoup.find_all('dl', class_='details lefty')

        # Extrae información de especie
        tomar_info = True
        for elemento_h2 in elementos_a:
            for elemento_a in elemento_h2.find_all('dd', class_='lng_ln'):
                if tomar_info:  # Solo procesar si la bandera está activa
                    elemento_p = elemento_a.find('b')
                    if elemento_p:
                        especie.append(elemento_p.text.strip())
                    else:
                        especie.append("null")
                    tomar_info = False  
                else:
                    tomar_info = True  

        # Extrae nombres
        elementos_a = ncbisoup.find_all('div', class_='rsltcont')
        for elemento_div in elementos_a:
            elemento_p = elemento_div.find('p', class_='title')
            if elemento_p:
                elemento_a = elemento_p.find('a')
                if elemento_a:
                    nombre.append(elemento_a.text.strip())

        # Extrae enlaces de descarga
        elementos_b = ncbisoup.find_all('div', class_='supp')
        for elemento_div in elementos_b:
            for elemento_a in elemento_div.find_all('a'):
                href = elemento_a.get('href')
                if href and href.startswith('/geo/download/?acc='):
                    descarga.append(href.strip())

        
        #descarga_file = driver.find_element(By.XPATH, "//a[contains(text(), 'Download data')]")
        #descarga_file.click()
        #elementos_b = ncbisoup.find_all('div', class_='section')
        #for elemento_div in elementos_b:
        #    for elemento_a in elemento_div.find_all('a'):




        # Busca el botón "Next" y haz clic en él
        try:
            siguiente_boton = driver.find_element(By.XPATH, "//a[contains(@class, 'active page_link next')]")
            siguiente_boton.click()  # Haz clic en el botón "Next"
            time.sleep(2)  # Esperar a que la nueva página cargue completamente
            contador_paginas += 1
        except Exception as e:
            print("No hay más páginas o error:", e)
            break  # Salir del bucle si no hay más páginas

    ##-----------------------------------------------------------------------------------------------------

    #for i in range(len(nombre)):
    #    hoja_activa.cell(row=i+1, column=1, value=nombre[i])
    #    hoja_activa.cell(row=i+1, column=2, value=especie[i])
    #    hoja_activa.cell(row=i+1, column=3, value=descarga[i])

    #Convertir datos en archivo .JSON
    #especie_json = json.dumps(especie)
    #nombre_json = json.dumps(nombre)
    #descarga_json = json.dumps(descarga)

    #Crear un DataFrame
    data = {
        'especie' : especie,
        'nombre' : nombre,
        'descarga' : descarga,
    }

    df = pd.DataFrame(data)

    # Convertir el DataFrame a formato JSON
    tabla = df.to_json(orient='records')
    #print(tabla)

    # Guardamos el libro de trabajo en un archivo
    #workbook.save("datos.xlsx")



    ##-----------------------------------------------------------------------------------------------------


    #igsr = requests.get('https://www.ncbi.nlm.nih.gov/gds/?term=human')
    #igsrSoup = BeautifulSoup(igsr.content, 'html.parser')


    ##-----------------------------------------------------------------------------------------------------

    return {tabla}

# #ebi = requests.get('https://www.ebi.ac.uk/ena/browser/text-search?query=Homo%20Sapiens')
#ebiSoup = BeautifulSoup(ebi.content, 'html.parser')


#ebinombre = []

#elementos_a = ncbisoup.find_all('div', class_='row')

#for elemento_div in elementos_a:
#    elemento_p = elemento_div.find('div', class_='col-xs-8')
#    if elemento_p:
#         nombre.append(elemento_p.text.strip())

#for i in range(len(nombre)):
#    hoja_activa.cell(row=i+1, column=1, value=nombre[i])
#    workbook.save("datosebi.xlsx")